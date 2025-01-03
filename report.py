import os
from fastapi import FastAPI, File, UploadFile, HTTPException, Request
from fastapi.responses import FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import (BarChart, LineChart, PieChart, DoughnutChart, Reference)
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import tempfile
import shutil

app = FastAPI()

# Use Render's or system's temporary directory
UPLOAD_DIR = os.environ.get("TMPDIR", tempfile.gettempdir())
app.mount("/static", StaticFiles(directory="."), name="static") 
templates = Jinja2Templates(directory=".") 

@app.get("/index")
async def get_index(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


def generate_report(input_file: BytesIO, month: str) -> str:
    wb = load_workbook(input_file)
    sheet = wb.active

    # Add titles, charts, and formatting
    add_report_title(sheet, month, sheet.max_column)
    create_charts(sheet, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)
    apply_currency_format(sheet, sheet.min_row, sheet.max_row, sheet.min_column, sheet.max_column)

    # Save report to temporary file
    report_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=UPLOAD_DIR)
    report_file_path = report_file.name
    wb.save(report_file_path)
    return report_file_path


def add_report_title(sheet, month, max_column):
    # Add a bold title for the report
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_column)
    sheet['K1'] = "Sales Report"
    sheet['K1'].font = Font(name='Calibri', bold=True, size=20)
    sheet['K1'].alignment = Alignment(horizontal="center")

    # Add the month below the title
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_column)
    sheet['K2'] = f"Month: {month}"
    sheet['K2'].font = Font(name='Calibri', bold=True, size=12)
    sheet['K2'].alignment = Alignment(horizontal="center")


def create_charts(sheet, min_row, max_row, min_column, max_column):
    # Set data for charts
    data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row)
    category = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row)

    # Bar Chart
    barchart = BarChart()
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(category)
    barchart.title = "Sales vs Product Line"
    barchart.x_axis.title = "Product Line"
    barchart.y_axis.title = "Sales"
    sheet.add_chart(barchart, f"{get_column_letter(max_column + 2)}{min_row}")

    # Line Chart
    linechart = LineChart()
    linechart.add_data(data, titles_from_data=True)
    linechart.set_categories(category)
    linechart.title = "Trend Analysis"
    linechart.x_axis.title = "Product Line"
    linechart.y_axis.title = "Sales"
    sheet.add_chart(linechart, f"{get_column_letter(max_column + 10)}{min_row}")

    # Pie Chart
    piechart = PieChart()
    piechart.add_data(data, titles_from_data=True)
    piechart.set_categories(category)
    piechart.title = "Product Line Distribution"
    sheet.add_chart(piechart, f"{get_column_letter(max_column + 2)}{max_row + 10}")

    # Doughnut Chart
    doughnut_chart = DoughnutChart()
    doughnut_chart.add_data(data, titles_from_data=False)
    doughnut_chart.set_categories(category)
    doughnut_chart.title = "Sales Distribution (Doughnut)"
    sheet.add_chart(doughnut_chart, f"{get_column_letter(max_column + 10)}{max_row + 10}")

def apply_currency_format(sheet, min_row, max_row, min_column, max_column):
    currency_style = NamedStyle(name="Currency", number_format="\u20B9#,##0.00")
    if "Currency" not in sheet.parent.named_styles:
        sheet.parent.add_named_style(currency_style)

    for i in range(min_column + 1, max_column + 1):
        letter = get_column_letter(i)
        sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
        sheet[f'{letter}{max_row + 1}'].style = currency_style

    sheet[f'{get_column_letter(min_column)}{max_row + 1}'] = 'Total'
    sheet[f'{get_column_letter(min_column)}{max_row + 1}'].font = Font(bold=True)



@app.post("/uploadfile/")
async def upload_file(file: UploadFile = File(...)):
    try:
        file_location = os.path.join(UPLOAD_DIR, file.filename)
        with open(file_location, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"message": f"File uploaded successfully: {file.filename}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")


@app.post("/generate-report/")
async def generate_report_endpoint(file: UploadFile = File(...), month: str = "January"):
    try:
        input_file = BytesIO(await file.read())
        report_file_path = generate_report(input_file, month)
        return FileResponse(report_file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            filename=os.path.basename(report_file_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating report: {str(e)}")


@app.get("/download-report/{report_file}")
async def download_report(report_file: str):
    try:
        file_path = os.path.join(UPLOAD_DIR, report_file)
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            filename=os.path.basename(file_path))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading report: {str(e)}")
